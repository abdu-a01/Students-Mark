from openpyxl import Workbook
from openpyxl.utils import get_column_letter as gcl
from .oneStud.utility.file_checker import checkFile,dict_xlsx
from .oneStud.grading import (
	gradeFun,
	gpaFun,
	pointFun,
	studNames,
	studMarks,
	subList,
	max_min
)


class StudMark:
	
	def __init__(self,file):
		self._studData = checkFile(file)
		self._students = studNames(self._studData)
		self._subjects = subList(self._studData)
		self._all_mark = studMarks(self._studData)
		self._total = len(self._students)
		self._max_sub,self._min_sub = max_min(self._studData)
		
	@property
	def studData(self):
		return self._studData
		
	@property
	def students(self):
		return self._students
	
	@property
	def subjects(self):
		return self._subjects
		
	@property
	def all_mark(self):
		return self._all_mark
	
	@property
	def total(self):
		return self._total
	
	@property
	def max_sub(self):
		return self._max_sub
		
	@property
	def min_sub(self):
		return self._min_sub
		
		
	def sort(self,gpa=False):
		data = self._studData.copy()
		
		keys = list(self._studData.keys())
		first = keys[0]
		
		if "gpa" in data[first] and gpa:
			data = {key:data[key] for key in sorted(data,key=lambda y: data[y]["gpa"],reverse=True)}
			
			return data
		
		data = {key:data[key] for key in sorted(data)}
		
		return data
		
		
	def sort_update(self,gpa=False):
		
		self._studData = self.sort(gpa=gpa).copy()
		
		
	
	def grade(self):
		data = self._studData.copy()
		
		subject = self._subjects.copy()
		
		for each in data:
			stud = data[each]
			for sub in subject:
				if "mark" not in stud[sub]:
					mark = stud[sub]
					grade_sub = gradeFun(mark)
				
					data[each][sub] = {
						"mark":mark,
						"grade":grade_sub
					}
					continue
				mark = stud[sub]["mark"]
				data[each][sub]["grade"] = gradeFun(mark)
		
		return data
	
	def grade_update(self):
		
		self._studData = self.grade().copy()
		
		
		
	def point(self):
		data = self._studData.copy()
		
		subject = self._subjects.copy()
		
		for each in data:
			stud = data[each]
			for sub in subject:
				if "mark" not in stud[sub]:
					mark = stud[sub]
					point_sub = pointFun(mark)
				
					data[each][sub] = {
						"mark":mark,
						"point":point_sub
					}
					continue
				mark = stud[sub]["mark"]
				data[each][sub]["point"] = pointFun(mark)
		
		return data
	
	
	def point_update(self):
		
		self._studData = self.point().copy()
		
		
		
	def gpa(self,cr_hours):
		data = self._studData.copy()
		marks = self._all_mark
		
		gpa_list = [gpaFun(each,cr_hours) for each in marks]
		
		for i,each in enumerate(self._students):
			data[each]["gpa"] = gpa_list[i]
		
		return data
	
	def gpa_update(self,cr_hours):
		self._studData = self.gpa(cr_hours).copy()
		
		
	def average(self):
		data = self._studData.copy()
		num_sub = len(self._subjects)
		mark_list = self._all_mark
		
		aver = [sum(marks)/num_sub for marks in mark_list]
		for i,each in enumerate(self._students):
			data[each]["average"] = aver[i]
		
		return data
		
	def average_update(self):
		self._studData = self.average().copy()
		
	def ranker(self,gpa=False):
		data = self._studData.copy()
		students = list(data.keys())
		
		if gpa and "gpa" in data[students[0]]:
			sorted_data = self.sort(gpa)
			
			stud_names = list(sorted_data.keys())
			
			rank = [(i,each) for i,each in enumerate(stud_names,1)]
			
		else:
			self.average_update()
			
			rank = [(i,each) for i,each in enumerate(sorted(data,key=lambda y:data[y]["average"],reverse=True),1)]
			
			self._studData = data.copy()
			
		
		for i,each in enumerate(rank):
			
			data[each]["rank"] = rank[i]
		
		return data
		
	def ranker_update(self,gpa=False):
		self._studData = self.ranker(gpa).copy()
			
	
	def id_giver(self,pre="",suf="",sep="/",length=4):
		data = self.sort()
		sep = sep if sep in ["-","\\",":"] else "/"
		form = lambda txt:f"{pre}{sep}{txt:}{sep}{suf}"
		
		IDs = [form(f"{i:0{length}d}") for i in range(1,len(data)+1)]
		
		for i,each in enumerate(data):
			data[each]["id_number"] = IDs[i]
		
		self._studData = data.copy()
	
	
	def one_sub(self,subject):
		data = self._studData.copy()
		marks = [data[name][subject] for name in data]
		
		class Subject:
			
			def max_mark(self, target=""):
				maximum = max(marks)
				result = {}
				for each in data:
					value = data[each][subject]
					if value == maximum:
						result[each] = value
				
				if target.lower().strip() == "sub":
					return maximum
				elif target.lower().strip() == "stud":
					studs = list(result.keys())
					return studs
				
				return result
				
				
			def min_mark(self, target=""):
				minimum = min(marks)
				result = {}
				for each in data:
					value = data[each][subject]
					if value == minimum:
						result[each] = value
				
				if target.lower().strip() == "Mark":
					return minimum
				elif target.lower().strip() == "stud":
					studs = list(result.keys())
					return studs
				
				return result
			
			def average(self):
				return sum(marks)/len(marks)
				
			def full_mark(self):
				result = {name:data[name][subject] for name in data}
				return result
				
		return Subject()
		
		
	def one_stud(self,name):
		data = self._studData.copy()
		stud_data = {name:data[name]}
		subjects = self._subjects.copy()
		
		class Student:
			
			def __init__(self):
				self.info = stud_data.copy()
				self.marks = [stud_data[name][sub] for sub in stud_data[name]]
			
			def max_mark(self,target=""):
				maximum = max(self.marks)
				result = {}
				for each in stud_data[name]:
					value = data[name][each]
					if value == maximum:
						result[each] = value
				target = target.lower().strip()
				if target == "mark":
					return maximum
				elif target == "sub":
					return list(result.keys())
				return result
				
				
			def min_mark(self,target=""):
				minimum = min(self.marks)
				result = {}
				for each in stud_data[name]:
					value = data[name][each]
					if value == minimum:
						result[each] = value
				target = target.lower().strip()
				if target == "mark":
					return minimum
				elif target == "sub":
					return list(result.keys())
				return result
			
			def average(self):
				return sum(self.marks)/len(self.marks)
			
			def grade(self,sub="all"):
				
				if sub in subjects:
					Mark = stud_data[name][sub]
					return gradeFun(Mark)
				
				grades = [gradeFun(each) for each in self.marks]
				
				return grades
			
			def point(self,sub="all"):
				
				if sub in subjects:
					Mark = stud_data[name][sub]
					return pointFun(Mark)
				
				point = [pointFun(each) for each in self.marks]
				
				return point
				
			def gpa(self,cr_hours):
				return gpaFun(self.marks,cr_hours)
			
		return Student()
		
	def json(self):
		json = __import__("json")
		with open("student_data.json","w") as file:
			json.dump(self._studData,file,indent=4)
		
		print("student data has been saved")
		
	def xlsx(self):
		data = self._studData
		rows = dict_xlsx(self._studData,self._students,self._subjects)
		wb = Workbook()

		ws = wb.active
		ws.title = "Student_data"
		
		for each in rows:
	
			ws.append(each)
		in_1 = self._students[0]
		
		subs = self._subjects
		
		mg_strt = 2
		for each in data[in_1]:
			each = data[in_1][each]
			if type(each) == dict:
				col_mv = len(each) - 1
				mg_end = mg_strt + len(each) - 1
				mv_strt = mg_strt + 1
				mv_end = mg_strt + len(subs) - 1
				ch1 = f"{gcl(mg_strt)}1"
				ch2 = f"{gcl(mg_end)}1"
				ch3 = f"{gcl(mv_strt)}1"
				ch4 = f"{gcl(mv_end)}1"
									
				ws.move_range(f"{ch3}:{ch4}",cols=col_mv)
				ws.merge_cells(f"{ch1}:{ch2}")
									
				mg_strt += len(each)
				
			
		
		wb.save("student_info.xlsx")
		
			
		